import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys
import os

# (name, email, password, pin, vin, vehicle, phone, country)
service_details = {'Demo Mode': [[1,1,"",""], (0,0,0,0,0,0,1,1)], 'Customer Enrollment': [[1,0,"Not finished",""], (1,1,1,1,0,1,1,1)], 'Add VIN': [[1,1,"Not finished",""], (1,1,1,1,1,1,1,1)],
            'App Registration Pages': [[1,1,"",""], (0,0,0,0,1,0,1,1)], 'App Log in-Log out': [[1,0,"",""], (0,1,1,0,0,1,1,1)], 'Nickname': [[1,1,"",""], (0,0,0,0,0,1,1,1)],
            'Services and licenses': [[1,1,"",""], (0,0,0,0,0,1,1,1)], 'Vehicle Status Report': [[0,0,"",""], (0,0,0,0,0,1,1,1)], 'Remote Lock-Unlock': [[0,0,""], (0,0,0,1,0,1,1,1)],
            'Remote Honk & Flash': [[0,0,"Not finished","chn"], (0,0,0,0,0,1,1,1)], 'My Car Statistics': [[1,1,"",""], (0,0,0,0,0,1,1,1)], 'My Cabin Comfort': [[0,0,"",""], (0,0,0,0,0,1,1,1)],
            'My Battery Charge': [[0,0,"",""], (0,0,0,0,0,1,1,1)], 'Service Management': [[1,1,"",""], (0,0,0,0,0,1,1,1)], 'Activate Heating': [[0,0,"","eur"], (0,0,0,0,0,1,1,1)],
            'Roadside Assistance': [[1,1,"",""], (0,0,0,0,0,1,1,1)], 'Data Services': [[1,1,"Broken",""], (0,0,0,0,0,1,1,1)], 'My Alerts': [[0,0,"Not finished","nar"], (0,0,0,0,0,1,1,1)],
            'Theft Alarm': [[0,0,"","eur"], (0,0,0,0,0,1,1,1)], 'Stolen Vehicle Locator': [[1,1,"Not finished","nar chn"], (0,0,0,0,0,1,1,1)], 'Audials': [[1,1,"","eur nar"], (0,0,0,0,0,1,1,1)],
            'Car Finder': [[0,1,"",""], (0,0,0,0,0,1,1,1)], 'Nav Companion': [[0,0,"",""], (0,0,0,0,0,1,1,1)], 'Notifications': [[1,0,"",""], (0,0,0,0,0,1,1,1)],
            'Push Notifications': [[0,1,"Broken",""], (0,0,0,0,0,1,1,1)], 'Profile': [[1,1,"",""], (1,1,0,1,0,1,1,1)], 'Localization': [[1,1,"Not automatable",""], (0,0,0,0,0,1,1,1)],
            'Privacy Mode': [[0,0,"",""], (1,1,1,1,0,1,1,1)], 'Remote Park Assist': [[0,0,"Not automatable",""], (1,1,1,1,0,1,1,1)], 'Stolen Vehicle Tracking': [[0,0,"","eur"], (0,0,0,0,0,1,1,1)]}
services = list(service_details.keys())

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def load_data():
    # source_path = resource_path("gui/Android_MY26_BY636_SQ_Remote_Services_EUR_Full - test.xlsx")
    source_path = resource_path("gui/Remote_Services.xlsx")
    workbook = load_workbook(source_path)

    skip_columns = ["A", "B", "D", "E", "G"]

    # Corrects the headers, swaps " for ' to avoid errors, replaces numeric list prefixes, removes sublists, removes empty lines,
    correct_headers = ['', 'TC ID', 'Region', 'Test Priority', 'Overall Effort (in Mins)', 'Test Case Title', 'Pre-Condition', 'Pre-Condition (Vehicle)', 'Action', 'Expected Result']
    for service in services:
        sheet = workbook[service]

        for col in sheet.iter_cols():
            col_letter = get_column_letter(col[0].column)
            if col_letter == "K":
                break
            if col_letter not in skip_columns:
                for cell in col:
                    if isinstance(cell.value, str):
                        cell.value = cell.value.replace("\"", "'")
                        while "\n\n" in cell.value:
                            cell.value = cell.value.replace("\n\n", "\n")

        col_num = 0
        for col in sheet.iter_cols():
            col_letter = get_column_letter(col[0].column)
            if col_letter == "K":
                break
            cell = sheet[f"{col_letter}4"]
            cell.value = correct_headers[col_num]

            col_num += 1

    workbook.save(source_path)

    TestCases = {}

    excel_file = pd.ExcelFile(source_path)
    sheet_names = excel_file.sheet_names
    for sheet in sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet, skiprows=3, usecols="B:J")
        test_case_list = []

        for _, row in df.iterrows():
            precon, action, expected = [], [], []
            unsplit_precon = row.get("Pre-Condition (Vehicle)")
            try:
                split_precon = unsplit_precon.split('\n')
                for txt in split_precon:
                    precon.append(txt)
            except:
                pass
            unsplit_action = row.get("Action")
            try:
                split_action = unsplit_action.split('\n')
                for txt in split_action:
                    action.append(txt)
            except:
                pass
            unsplit_expected = row.get("Expected Result")
            try:
                split_expected = unsplit_expected.split('\n')
                for txt in split_expected:
                    expected.append(txt)
            except:
                pass
            test_case = {
                "Region": row.get("Region", "").strip(),
                "Test Case Description": row.get("Test Case Title", "").strip(),
                "Pre-Condition": precon,
                "Action": action,
                "Expected Result": expected
            }
            test_case_list.append(test_case)

        TestCases[sheet] = test_case_list

    return TestCases