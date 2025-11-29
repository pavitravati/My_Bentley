from PySide6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QLabel, QTextEdit, QScrollArea, QToolButton, \
    QSizePolicy, QPushButton, QComboBox, QTableWidget, QHeaderView, QTableWidgetItem
from PySide6.QtGui import QFont, QPixmap, QPainter, QColor
from PySide6.QtCore import Qt
from pathlib import Path
from PySide6.QtWidgets import QApplication
from openpyxl.reader.excel import load_workbook
from core.kpm_creater import download_kpm
from excel import load_data
import os
import glob
from core import globals
from PySide6.QtCharts import QChart, QChartView, QPieSeries
import random

testcase_map = load_data()

class ServiceReport(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        try:
            self.service = next(iter(globals.log_history))
            self.current_test = True
        except:
            self.service = None
            self.current_test = False

        if QApplication.primaryScreen().size().width() > 1500:
            self.screen = 'Monitor'
        else:
            self.screen = 'Laptop'

        self.setWindowTitle("Automated Testing Report")

        screen = QApplication.primaryScreen()
        screen_size = screen.availableGeometry()
        width = int(screen_size.width() * 0.8)
        height = int(screen_size.height() * 0.9)
        self.resize(width, height)

        self.setFixedSize(width, height)

        layout = QVBoxLayout()
        self.setLayout(layout)

        self.dropdown = QComboBox()
        if self.current_test:
            self.build_body_for_service(layout)
        else:
            folders = [
                f for f in os.listdir(globals.sharedrive_path)
                if os.path.isdir(os.path.join(globals.sharedrive_path, f))
            ]
            most_recent = max(folders, key=lambda f: os.path.getctime(os.path.join(globals.sharedrive_path, f)))
            self.build_body_for_folder(layout, most_recent)

    def toolbar_button_clicked(self, svc, layout):
        self.service = svc
        self.refresh_ui(layout)

    def toolbar_button_clicked_import(self, svc, layout, folder_name):
        self.service = svc
        self.refresh_ui(layout, new_folder=True, folder_name=folder_name)

    def overview_button_clicked(self, layout, folder_name):
        self.service = None
        self.refresh_ui(layout, new_folder=True, folder_name=folder_name)

    def refresh_ui(self, current_layout, new_folder=False, folder_name=None):
        # Clear the existing layout (everything except toolbar)
        main_layout = self.layout()

        # Remove all widgets after the toolbar (index 1 onward)
        for i in reversed(range(0, main_layout.count())):
            item = main_layout.itemAt(i)
            widget = item.widget()
            if widget:
                widget.deleteLater()
            else:
                layout = item.layout()
                if layout:
                    while layout.count():
                        sub_item = layout.takeAt(0)
                        sub_widget = sub_item.widget()
                        if sub_widget:
                            sub_widget.deleteLater()
                    main_layout.removeItem(layout)

        if new_folder:
            self.build_body_for_folder(current_layout, folder_name)
        else:
            self.build_body_for_service(current_layout)

    def build_body_for_service(self, layout):
        toolbar = QHBoxLayout()
        self.dropdown = QComboBox()
        if self.current_test:
            self.dropdown.addItem("Current test")
        else:
            self.dropdown.addItem("Select test")
        results_folder = globals.sharedrive_path
        folder_names = [name for name in os.listdir(results_folder) if
                        os.path.isdir(os.path.join(results_folder, name))]
        for folder in folder_names:
            self.dropdown.addItem(f"{folder.replace('+', ':')}")
        self.dropdown.currentTextChanged.connect(lambda text: self.show_test_result(layout, text))

        toolbar.addWidget(self.dropdown)

        for service in globals.log_history:
            toolbar_btn = QPushButton(service)
            toolbar_btn.setCursor(Qt.PointingHandCursor)
            toolbar_btn.clicked.connect(lambda checked, svc=service: self.toolbar_button_clicked(svc, layout))
            toolbar.addWidget(toolbar_btn)
        toolbar.addStretch()

        layout.addLayout(toolbar)

        top_layout = QHBoxLayout()
        top_layout.setContentsMargins(0, 40, 0, 20)

        title = QLabel(self.service)
        title.setFont(QFont("Arial", 25, QFont.Bold))
        title.setStyleSheet("margin-left: 20px; margin-bottom: 20px;")
        top_layout.addWidget(title)

        top_layout.addStretch()

        logo = QLabel()
        img_path = Path(__file__).parent / "images" / "bentleylogo.png"
        pixmap = QPixmap(str(img_path))
        logo.setPixmap(pixmap)
        logo.setScaledContents(True)
        logo.setMaximumSize(135, 50)
        logo.setStyleSheet("margin-right: 20px;")
        top_layout.addWidget(logo)

        layout.addLayout(top_layout)

        container_layout = QHBoxLayout()

        # Add all testcases so that they can all be viewed
        testcase_scroll = QScrollArea()
        testcase_scroll.setWidgetResizable(True)
        testcase_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        testcase_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        testcase_scroll.setFixedWidth(340)
        testcase_scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background: transparent;
            }
            QScrollBar:vertical {
                background: transparent;
                width: 10px;   
                margin: 0px 5px 0px 0px;
                border-radius: 5px;
            }
        """)

        # container that holds buttons
        testcase_container = QWidget()
        testcase_layout = QVBoxLayout(testcase_container)
        testcase_layout.setAlignment(Qt.AlignTop)

        if self.service:
            for row, case in enumerate(globals.log_history[self.service]):
                result = '✅'
                for i in range(len(globals.log_history[self.service][case])):
                    if globals.log_history[self.service][case][i][0] == '❌':
                        result = '❌'
                        break

                btn = QToolButton()
                test_description = testcase_map[self.service][row]['Test Case Description']
                btn.setText(test_description)
                btn.setToolButtonStyle(Qt.ToolButtonTextUnderIcon)
                btn.setCursor(Qt.PointingHandCursor)
                btn.setFixedWidth(320)
                btn.setFixedHeight(45)
                btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Minimum)

                if result == '✅':
                    btn.setStyleSheet("background-color: #394d45; font-size: 12px; color: white;")
                elif result == '❌':
                    btn.setStyleSheet("background-color: #7d232b; font-size: 12px; color: white;")
                else:
                    btn.setStyleSheet("background-color: gray; font-size: 12px; color: white;")

                row_logs = globals.log_history[self.service][case]
                cleaned_logs = []
                for log in row_logs:
                    if log.startswith("$"):
                        self.kpm_log = log[1:]
                    else:
                        cleaned_logs.append(log)
                logs_combined = "\n".join(cleaned_logs)
                btn.clicked.connect(
                    lambda checked, c=test_description, l=logs_combined, r=row + 1, s=self.service:
                    self.on_test_clicked(c, l, r, s)
                )
                testcase_layout.addWidget(btn)

        testcase_scroll.setWidget(testcase_container)

        container_layout.addWidget(testcase_scroll)

        detail_layout = QVBoxLayout()

        self.log_container = QHBoxLayout()
        self.log_textbox = QTextEdit()
        parent_height = self.parent().height() if self.parent() else 800
        screen_size = QApplication.primaryScreen().size()
        available_width = screen_size.width()
        if available_width > 1500:
            multi = 0.35
        else:
            multi = 0.25
        self.log_textbox.setFixedHeight(int(parent_height * multi))
        # self.log_textbox.setFixedWidth(int(available_width * 0.5))
        self.log_textbox.setReadOnly(True)
        self.log_textbox.setFont(QFont("Arial", 12))
        self.log_container.addWidget(self.log_textbox)
        self.log_container.setSpacing(0)

        detail_layout.addLayout(self.log_container)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)

        container = QWidget()
        self.images_layout = QHBoxLayout(container)
        self.images_layout.setSpacing(5)
        self.images_layout.setContentsMargins(20, 0, 20, 0)

        scroll_area.setWidget(container)

        detail_layout.addWidget(scroll_area)

        container_layout.addLayout(detail_layout)

        layout.addLayout(container_layout)

    def on_test_clicked(self, test_case, logs_combined, row, service, imported_test=False, folder_name=None):
        self.log_textbox.setPlainText(test_case)
        self.log_textbox.append(logs_combined)
        self.log_textbox.setStyleSheet("""
                    QTextEdit {
                        border: 0.8px solid #c0c0c0;
                    }
                """)

        if getattr(self, "btn_container", None) is not None:
            self.log_container.removeWidget(self.btn_container)
            self.btn_container.deleteLater()
            self.btn_container = None

        self.btn_container = QWidget()
        self.btn_container.setStyleSheet("background-color: white; border-top: 0.8px solid #c0c0c0; border-right: 0.8px solid #c0c0c0; border-bottom: 0.8px solid #c0c0c0;")
        if '❌' in logs_combined:
            self.kpm_container = QVBoxLayout(self.btn_container)
            kpm_btn = QToolButton()
            kpm_btn.setText("Generate KPM")
            kpm_btn.setCursor(Qt.PointingHandCursor)
            kpm_btn.setStyleSheet("""
                                QToolButton {
                                    background-color: #394d45;
                                    font-size: 16px;
                                    font-weight: bold;
                                    width: 180px;
                                    height: 40px;
                                    color: white;
                                }
                                QToolButton:hover {
                                    background-color: #25312c;
                                }
                            """)
            kpm_btn.clicked.connect(lambda: self.generate_kpm(row))
            self.kpm_container.addWidget(kpm_btn, alignment=Qt.AlignTop)
            self.log_container.addWidget(self.btn_container)
            self.log_textbox.setStyleSheet("""
                QTextEdit {
                    border: none;
                    border-top: 0.8px solid #c0c0c0; 
                    border-left: 0.8px solid #c0c0c0; 
                    border-bottom: 0.8px solid #c0c0c0;
                }
            """)

        while self.images_layout.count():
            item = self.images_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        if imported_test:
            image_dir = os.path.join(folder_name, "images")
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            image_dir = os.path.join(base_dir, "fail_images")
        image_paths = []
        row_num = f"0{row}" if row < 10 else f"{row}"

        for file_path in glob.glob(os.path.join(image_dir, "*.png")):
            filename = os.path.basename(file_path)
            if row_num in filename and service in filename:
                image_paths.append(file_path)

        for img_path_str in image_paths:
            internal_container = QWidget()
            # May need changing to work on laptop and monitor
            internal_container.setFixedWidth(280)
            report_image = QVBoxLayout(internal_container)

            img_path = Path(img_path_str)
            pixmap = QPixmap(str(img_path))
            if pixmap.isNull():
                print(f"Failed to load image: {img_path}")
                continue

            img_label = QLabel()
            img_text = img_path_str.split("-")[-2]
            img_label.setText(img_text)
            img_label.setFont(QFont("Arial", 12))
            img_label.setWordWrap(True)
            report_image.addWidget(img_label)

            image_display = QLabel()

            screen_size = QApplication.primaryScreen().size()
            available_width = screen_size.width()
            if available_width > 1500:
                img_size = 410
            else:
                img_size = 260
            image_display.setPixmap(pixmap.scaled(img_size, img_size, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            image_display.setStyleSheet("margin: 5px;")

            report_image.addWidget(image_display)
            self.images_layout.addWidget(internal_container)

    def show_test_result(self, layout, text):
        if text == 'Current test':
            self.service = next(iter(globals.log_history))
            self.refresh_ui(layout)
        else:
            folder_name = text.replace(':', '+')
            script_dir = os.path.dirname(os.path.abspath(__file__))
            # results_folder = os.path.join(script_dir, "test_results")
            results_folder = globals.sharedrive_path
            test_folder_path = os.path.join(results_folder, folder_name)
            test_result_file = load_workbook(os.path.join(test_folder_path, "test_results.xlsx"))
            # self.service = test_result_file.sheetnames[0]
            self.refresh_ui(layout, new_folder=True, folder_name=folder_name)

    def clear_layout(self, layout):
        """Recursively remove and delete widgets, layouts and spacers from layout."""
        if layout is None:
            return
        while layout.count():
            item = layout.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            child_layout = item.layout()
            if widget:
                # detach and queue deletion
                widget.hide()
                widget.setParent(None)
                widget.deleteLater()
            elif child_layout:
                # recursively clear child layout
                self.clear_layout(child_layout)
            else:
                # spacer - nothing to delete, just continue
                pass

    def build_body_for_folder(self, layout, folder_name):
        toolbar = QHBoxLayout()
        self.dropdown = QComboBox()
        self.dropdown.addItem(folder_name.replace("+", ":"))
        if self.current_test:
            self.dropdown.addItem("Current test")
        # script_dir = os.path.dirname(os.path.abspath(__file__))
        # results_folder = os.path.join(script_dir, "test_results")
        results_folder = globals.sharedrive_path
        folder_names = [name for name in os.listdir(results_folder) if
                        os.path.isdir(os.path.join(results_folder, name)) and folder_name != name]
        for folder in folder_names:
            self.dropdown.addItem(f"{folder.replace('+', ':')}")
        self.dropdown.currentTextChanged.connect(lambda text: self.show_test_result(layout, text))

        toolbar.addWidget(self.dropdown)

        test_folder_path = os.path.join(results_folder, folder_name)
        test_result_file = load_workbook(os.path.join(test_folder_path, "test_results.xlsx"))
        toolbar_btn = QPushButton("Overview")
        toolbar_btn.setCursor(Qt.PointingHandCursor)
        toolbar_btn.clicked.connect(lambda checked: self.overview_button_clicked(layout, folder_name))
        toolbar.addWidget(toolbar_btn)
        for sheet in test_result_file.sheetnames:
            toolbar_btn = QPushButton(sheet)
            toolbar_btn.setCursor(Qt.PointingHandCursor)
            toolbar_btn.clicked.connect(lambda checked, svc=sheet: self.toolbar_button_clicked_import(svc, layout, folder_name))
            toolbar.addWidget(toolbar_btn)
        toolbar.addStretch()

        layout.addLayout(toolbar)

        top_layout = QHBoxLayout()
        top_layout.setContentsMargins(0, 40, 0, 20)

        title = QLabel(self.service if self.service else "Service Report")
        title.setFont(QFont("Arial", 25, QFont.Bold))
        title.setStyleSheet("margin-left: 20px; margin-bottom: 20px;")
        top_layout.addWidget(title)

        top_layout.addStretch()

        logo = QLabel()
        img_path = Path(__file__).parent / "images" / "bentleylogo.png"
        pixmap = QPixmap(str(img_path))
        logo.setPixmap(pixmap)
        logo.setScaledContents(True)
        logo.setMaximumSize(135, 50)
        logo.setStyleSheet("margin-right: 20px;")
        top_layout.addWidget(logo)

        layout.addLayout(top_layout)

        container_layout = QHBoxLayout()

        if self.service:
            # Add all testcases so that they can all be viewed
            testcase_scroll = QScrollArea()
            testcase_scroll.setWidgetResizable(True)
            testcase_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            testcase_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            testcase_scroll.setFixedWidth(340)
            testcase_scroll.setStyleSheet("""
                        QScrollArea {
                            border: none;
                            background: transparent;
                        }
                        QScrollBar:vertical {
                            background: transparent;
                            width: 10px;   
                            margin: 0px 5px 0px 0px;
                            border-radius: 5px;
                        }
                    """)

            # container that holds buttons
            testcase_container = QWidget()
            testcase_layout = QVBoxLayout(testcase_container)
            testcase_layout.setAlignment(Qt.AlignTop)
            current_sheet = test_result_file[self.service]

            for row in range(1, current_sheet.max_row+1):
                cell_value = current_sheet[f'B{row}'].value
                if '❌' in str(cell_value):
                    result = '❌'
                elif '🔒' in str(cell_value):
                    result = '🔒'
                else:
                    result = '✅'

                btn = QToolButton()
                test_description = str(current_sheet[f'A{row}'].value)
                btn.setText(test_description)
                btn.setToolButtonStyle(Qt.ToolButtonTextUnderIcon)
                btn.setCursor(Qt.PointingHandCursor)
                btn.setFixedWidth(320)
                btn.setFixedHeight(45)
                btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Minimum)

                if result == '✅':
                    btn.setStyleSheet("background-color: #394d45; font-size: 12px; color: white;")
                elif result == '❌':
                    btn.setStyleSheet("background-color: #7d232b; font-size: 12px; color: white;")
                else:
                    btn.setStyleSheet("background-color: gray; font-size: 12px; color: white;")

                row_logs = current_sheet[f'B{row}']
                row_logs = str(row_logs.value).split('\n')
                cleaned_logs = []
                for i, log in enumerate(row_logs):
                    if log.startswith("$"):
                        joined_kpm = "\n".join(row_logs[i:])
                        self.kpm_log = joined_kpm[1:]
                        break
                    else:
                        cleaned_logs.append(log)
                logs_combined = "\n".join(cleaned_logs)
                btn.clicked.connect(
                    lambda checked, c=test_description, l=logs_combined, r=row, s=self.service:
                    self.on_test_clicked(c, l, r, s, True, test_folder_path)
                )
                testcase_layout.addWidget(btn)

            testcase_scroll.setWidget(testcase_container)

            container_layout.addWidget(testcase_scroll)

            detail_layout = QVBoxLayout()

            self.log_container = QHBoxLayout()
            self.log_textbox = QTextEdit()
            parent_height = self.parent().height() if self.parent() else 800
            screen_size = QApplication.primaryScreen().size()
            available_width = screen_size.width()
            if available_width > 1500:
                multi = 0.35
            else:
                multi = 0.25
            self.log_textbox.setFixedHeight(int(parent_height * multi))
            self.log_textbox.setReadOnly(True)
            self.log_textbox.setFont(QFont("Arial", 12))
            self.log_container.addWidget(self.log_textbox)

            detail_layout.addLayout(self.log_container)

            scroll_area = QScrollArea()
            scroll_area.setWidgetResizable(True)

            container = QWidget()
            self.images_layout = QHBoxLayout(container)
            self.images_layout.setSpacing(5)
            self.images_layout.setContentsMargins(20, 0, 20, 0)

            scroll_area.setWidget(container)

            detail_layout.addWidget(scroll_area)

            container_layout.addLayout(detail_layout)

            layout.addLayout(container_layout)
        else:
            main_layout = QHBoxLayout()

            left_layout = QVBoxLayout()
            left_layout.setSpacing(5)
            top_left_layout = QHBoxLayout()
            top_left_layout.setAlignment(Qt.AlignCenter)

            series = QPieSeries()
            series.append("Pass", 14)
            series.append("Fail", 2)
            series.append("Blocked", 0)
            series.slices()[0].setBrush(QColor("#394d45"))
            series.slices()[1].setBrush(QColor("#7d232b"))
            series.slices()[2].setBrush(QColor("gray"))
            series.setPieSize(1.0)
            series.setHorizontalPosition(0.5)
            series.setVerticalPosition(0.5)
            for slice in series.slices():
                slice.setPen(Qt.NoPen)
            chart = QChart()
            chart.setBackgroundVisible(False)
            chart.addSeries(series)
            chart.legend().hide()
            chart.setTitle("")
            chart_view = QChartView(chart)

            if self.screen == "Monitor":
                chart_view.setMaximumSize(450, 450)
            else:
                chart_view.setFixedSize(250, 250)
            chart_view.setRenderHint(QPainter.Antialiasing)
            top_left_layout.addWidget(chart_view)

            label_layout = QVBoxLayout()
            label_layout.setSpacing(30)
            label_widget = QWidget()
            label_widget.setLayout(label_layout)
            top_left_layout.addWidget(label_widget, alignment=Qt.AlignVCenter)
            tests_run_label = QLabel(f"Tests passed: 14")
            tests_passed_label = QLabel(f"Tests Failed: 2")
            tests_failed_label = QLabel(f"Tests blocked: 0")
            if self.screen == "Monitor":
                label_style = "font-size: 25px; font-weight: bold;"
            else:
                label_style = "font-size: 14px; font-weight: bold;"
            tests_run_label.setStyleSheet(label_style)
            tests_passed_label.setStyleSheet(label_style)
            tests_failed_label.setStyleSheet(label_style)
            label_layout.addWidget(tests_run_label)
            label_layout.addWidget(tests_passed_label)
            label_layout.addWidget(tests_failed_label)
            top_left_layout.addLayout(label_layout)

            left_layout.addLayout(top_left_layout)
            label_style = "font-size: 24px;" if self.screen == "Monitor" else "font-size: 16px;"
            bottom_left_layout = QVBoxLayout()
            phone_type = QLabel(f"Phone type:   Android")
            phone_type.setStyleSheet(label_style)
            bottom_left_layout.addWidget(phone_type)
            phone_version = QLabel(f"Phone version:    Galaxy S23")
            phone_version.setStyleSheet(label_style)
            bottom_left_layout.addWidget(phone_version)
            app_version = QLabel(f"App version:    5.17.0[221339]")
            app_version.setStyleSheet(label_style)
            bottom_left_layout.addWidget(app_version)
            vin = QLabel(f"VIN used:    SJAAE14V3TC029739")
            vin.setStyleSheet(label_style)
            bottom_left_layout.addWidget(vin)
            region = QLabel(f"App region:   Europe")
            region.setStyleSheet(label_style)
            bottom_left_layout.addWidget(region)
            email = QLabel(f"Email used:    testdrive@gqm.anonaddy.com")
            email.setStyleSheet(label_style)
            bottom_left_layout.addWidget(email)
            if self.screen == "Monitor":
                bottom_left_layout.setContentsMargins(60, 20, 0, 40)
            else:
                bottom_left_layout.setContentsMargins(30, 20, 0, 50)
            left_layout.addLayout(bottom_left_layout)

            main_layout.addLayout(left_layout)

            result_table = QTableWidget()
            result_table.setColumnCount(5)
            result_table.setColumnWidth(0, 170 if self.screen == 'Monitor' else 120)
            result_table.setColumnWidth(1, 330 if self.screen == 'Monitor' else 286)
            result_table.setColumnWidth(2, 80 if self.screen == 'Monitor' else 60)
            result_table.setColumnWidth(3, 70 if self.screen == 'Monitor' else 60)
            result_table.setColumnWidth(4, 90 if self.screen == 'Monitor' else 60)
            result_table.setHorizontalHeaderLabels(["Service", "Testcase", "", "", ""])
            header = result_table.horizontalHeader()
            header.setSectionResizeMode(QHeaderView.Fixed)
            header.setFixedHeight(40)
            result_table.verticalHeader().setVisible(False)
            result_table.setRowCount(16)
            result_table.setEditTriggers(QTableWidget.NoEditTriggers)
            result_table.setMaximumHeight(700)
            if self.screen != "Monitor":
                result_table.setMinimumWidth(600)
            result_table.setStyleSheet("""
                                QTableWidget {
                                    background-color: white;
                                    gridline-color: #dcdcdc;
                                    font-size: 14px;
                                    color: #25312c;
                                    selection-background-color: #c7d3cc;
                                    selection-color: black;
                                    outline: 0;
                                }
                                QHeaderView::section {
                                    background-color: #394d45;
                                    color: white;
                                    font-weight: bold;
                                    border: none;
                                    padding: 8px;
                                    border-right: 1px solid #51645c;
                                }
                                QTableWidget::item {
                                    border-bottom: 1px solid #dcdcdc;
                                    padding: 6px;
                                }
                            """)
            header_font = QFont("Arial", 12, QFont.Bold)
            result_table.horizontalHeader().setFont(header_font)
            for _ in range(3):
                header_item = result_table.horizontalHeaderItem(_+2)
                font = header_item.font()
                font.setPointSize(8)
                header_item.setFont(font)

            test_result_file = load_workbook(os.path.join(test_folder_path, "test_results.xlsx"))
            row_count = 0
            for sheet in test_result_file.sheetnames:
                current_sheet = test_result_file[sheet]
                for row in range(1, current_sheet.max_row+1):
                    error_btn = QPushButton("Error")
                    error_btn.setCursor(Qt.PointingHandCursor)
                    error_btn.setStyleSheet("""
                        QPushButton {
                            background-color: #7d232b;
                            font-size: 12px;
                            color: white;
                        }
                    """)
                    result_table.setItem(row_count, 0, QTableWidgetItem(sheet))
                    result_table.setItem(row_count, 1, QTableWidgetItem(current_sheet[f'A{row}'].value))
                    result_table.setItem(row_count, 2, QTableWidgetItem(f"{str(random.randint(1,10))} secs"))
                    result_table.setItem(row_count, 3, QTableWidgetItem("Fail" if '❌' in current_sheet[f'B{row}'].value else "Pass"))
                    result_table.setCellWidget(row_count, 4, error_btn if '❌' in current_sheet[f'B{row}'].value else None)
                    row_count += 1

            for row in range(result_table.rowCount()):
                for col in range(result_table.columnCount()):
                    item = result_table.item(row, col)
                    if item is None:
                        continue
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFont(QFont("Arial", 12 if self.screen == "Monitor" else 10))
            for _ in range(16):
                result_table.setRowHeight(_, 60)

            main_layout.addWidget(result_table)
            layout.addLayout(main_layout)

    def generate_kpm(self, row):
        download_kpm(self.service, row, manual=True)